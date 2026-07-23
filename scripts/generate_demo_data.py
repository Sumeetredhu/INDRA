"""Build the synthetic plant corpus that the whole INDRA demo stands on.

Ten documents plus two rendered assets, written into ``settings.demo_dir``, from the single fact
sheet in :mod:`scripts.demo_facts`. Nothing here types a threshold, a date or a name of its own:
every number that appears in more than one file is imported, so the cross-document claims the demo
makes ("78% measured against an 85% OEM limit", "wear pattern similar to the 2022 failure that cost
14 hours and Rs 25 lakh") are true by construction rather than by proofreading.

What is produced
----------------

======================================  ===========================================================
File                                     Why the demo needs it
======================================  ===========================================================
``P-101_OEM_Manual.pdf``                 The authority for every threshold. 85% bearing wear limit.
``WO_2024_0342_P-101.pdf``               Open work order, 78% wear, handwritten margin annotation.
``Inspection_2024_0315_P-101.pdf``       The sentence that bridges 2024 condition to 2022 failure.
``ShiftLog_2024_0614.pdf``               Two alarm bypasses nobody connected to the bearing.
``Incident_2022_0820_P-101.pdf``         The business case: 14 h, Rs 25,00,000.
``RCA_2022_0825_P-101.pdf``              The mechanism: lubrication failure via LP-101A.
``P-101_P&ID.png``                       A real raster drawing the vision parser genuinely parses.
``P-101_P&ID_scanned.png``               The same drawing degraded, so OCR tag correction is real.
``Email_Retirement_Rajesh.pdf``          The knowledge cliff, hiding in an HR e-mail.
``SOP_Bearing_Replacement.pdf``          12 steps, 240 minutes, five hold points.
``Factory_Act_Section41b.pdf``           The obligation the compliance audit measures V-201 against.
``Maintenance_Log_2022_2024.xlsx``       Structured history behind every claim the prose makes.
``photos/P-101_nameplate.jpg``           Input for the photo-to-query beat (not part of the corpus).
======================================  ===========================================================

Determinism
-----------

Regenerating on any machine must produce the same bytes, because ingestion is content-addressed
(``docs/DECISIONS.md`` D6) and a demo rehearsal that duplicates every node ruins the graph. So:

* every date comes from :data:`scripts.demo_facts.REFERENCE_DATE` and its siblings — never
  ``date.today()``;
* ReportLab runs with ``rl_config.invariant``, which fixes the PDF creation date and document id;
* the ``.xlsx`` container is rewritten with fixed ZIP entry timestamps, because ``openpyxl`` stamps
  wall-clock time into every member;
* every random draw (handwriting jitter, scan noise, nameplate grain) is seeded from
  ``settings.llm_seed`` plus a fixed per-artefact offset.

Run it::

    python -m scripts.generate_demo_data            # into data/demo
    python -m scripts.generate_demo_data --output /tmp/corpus --seed 7
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import sys
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Final, Sequence

if __package__ in (None, ""):  # pragma: no cover - allows `python scripts/generate_demo_data.py`
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from indra.core.config import Settings, get_settings
from indra.core.exceptions import BlobStoreError, ConfigurationError, IndraError
from indra.core.logging import get_logger

from scripts import demo_facts as facts
from scripts._console import Console
from scripts.demo_drawings import (
    PID_CONNECTIONS,
    PID_HEIGHT,
    PID_SYMBOLS,
    PID_WIDTH,
    degrade_scan,
    render_handwriting,
    render_nameplate,
    render_pid,
    save_image,
)

logger = get_logger(__name__)

MANIFEST_FILENAME: Final[str] = "manifest.json"
"""Published alongside the corpus. ``seed_demo_data`` and ``run_demo_check`` read it rather than
re-deriving constants, so there is exactly one place a fact can be wrong."""

#: Per-artefact seed offsets. Fixed so that changing one artefact's look cannot silently change
#: another's; added to ``settings.llm_seed`` to get the actual seed.
_SEED_OFFSET_SCAN: Final[int] = 101
_SEED_OFFSET_HANDWRITING: Final[int] = 202
_SEED_OFFSET_NAMEPLATE: Final[int] = 303

#: ZIP entry timestamp for the spreadsheet container, so the file is byte-stable.
_ZIP_TIMESTAMP: Final[tuple[int, int, int, int, int, int]] = facts.FIXED_TIMESTAMP
_FIXED_DATETIME: Final[datetime] = datetime(*facts.FIXED_TIMESTAMP)

# --------------------------------------------------------------------------------------
# Page geometry (mm). A4 portrait with an 18 mm margin all round.
# --------------------------------------------------------------------------------------
_MARGIN_MM: Final[float] = 18.0
_CONTENT_WIDTH_MM: Final[float] = 210.0 - 2 * _MARGIN_MM


# ======================================================================================
# Reporting types
# ======================================================================================


@dataclass(frozen=True, slots=True)
class GeneratedFile:
    """One artefact written to disk, with the identity ingestion will address it by."""

    filename: str
    path: Path
    size_bytes: int
    sha256: str
    document_type: str
    title: str
    role: str
    pages: int | None = None

    def as_manifest_entry(self) -> dict[str, Any]:
        spec = facts.document_spec(self.filename)
        return {
            "filename": self.filename,
            "size_bytes": self.size_bytes,
            "sha256": self.sha256,
            "document_type": self.document_type,
            "title": self.title,
            "role": self.role,
            "pages": self.pages,
            "document_date": spec.document_date.isoformat() if spec else None,
            "equipment_tags": list(spec.equipment_tags) if spec else [],
            "people": list(spec.people) if spec else [],
            "key_facts": list(spec.key_facts) if spec else [],
            "description": spec.description if spec else "",
        }


@dataclass(frozen=True, slots=True)
class CorpusReport:
    """The outcome of one generation run."""

    output_dir: Path
    files: tuple[GeneratedFile, ...]
    manifest_path: Path
    seed: int
    reference_date: date
    warnings: tuple[str, ...] = ()

    @property
    def total_bytes(self) -> int:
        return sum(item.size_bytes for item in self.files)

    @property
    def corpus_files(self) -> tuple[GeneratedFile, ...]:
        """The files that get ingested, excluding assets such as the nameplate photo."""
        return tuple(item for item in self.files if item.role == "corpus")

    def by_name(self, filename: str) -> GeneratedFile | None:
        for item in self.files:
            if item.filename == filename:
                return item
        return None


# ======================================================================================
# ReportLab plumbing
# ======================================================================================


@dataclass(frozen=True, slots=True)
class PageFurniture:
    """The running header and footer of one generated PDF."""

    organisation: str
    title: str
    document_number: str
    document_date: date
    classification: str = "INTERNAL — PLANT RECORDS"


def _reportlab() -> Any:
    """Import ReportLab lazily and turn its absence into an actionable configuration error."""
    try:
        import reportlab  # noqa: F401
        import reportlab.rl_config as rl_config
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.platypus import (
            HRFlowable,
            Image as RLImage,
            KeepTogether,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover - dependency is declared in requirements.txt
        raise ConfigurationError(
            "ReportLab is required to generate the demo corpus PDFs. Install it with "
            "`pip install reportlab` (it is already listed in requirements.txt).",
            cause=exc,
        ) from exc

    # Fixed PDF creation date and document id. Without this every regeneration produces different
    # bytes, a different SHA-256, and therefore a duplicate document set on every rehearsal (D6).
    rl_config.invariant = 1

    return {
        "colors": colors, "A4": A4, "mm": mm, "ParagraphStyle": ParagraphStyle,
        "TA_CENTER": TA_CENTER, "TA_JUSTIFY": TA_JUSTIFY, "TA_LEFT": TA_LEFT,
        "pdfmetrics": pdfmetrics, "canvas": pdf_canvas, "HRFlowable": HRFlowable,
        "Image": RLImage, "KeepTogether": KeepTogether, "PageBreak": PageBreak,
        "Paragraph": Paragraph, "SimpleDocTemplate": SimpleDocTemplate, "Spacer": Spacer,
        "Table": Table, "TableStyle": TableStyle,
    }


def _styles(rl: dict[str, Any]) -> dict[str, Any]:
    """Paragraph styles for the generated documents.

    Built fresh per document build rather than shared at module scope: ReportLab styles are mutable
    and a leaked mutation between documents would be invisible until a table looked wrong.
    """
    ps = rl["ParagraphStyle"]
    colors = rl["colors"]
    ink = colors.HexColor("#111111")
    slate = colors.HexColor("#40474f")

    base = ps("body", fontName="Helvetica", fontSize=9.2, leading=13.0, textColor=ink,
              alignment=rl["TA_JUSTIFY"], spaceAfter=5)
    return {
        "title": ps("title", parent=base, fontName="Helvetica-Bold", fontSize=16, leading=19,
                    alignment=rl["TA_LEFT"], spaceAfter=4),
        "subtitle": ps("subtitle", parent=base, fontName="Helvetica", fontSize=10, leading=13,
                       textColor=slate, alignment=rl["TA_LEFT"], spaceAfter=10),
        "h1": ps("h1", parent=base, fontName="Helvetica-Bold", fontSize=11.5, leading=15,
                 alignment=rl["TA_LEFT"], spaceBefore=11, spaceAfter=5),
        "h2": ps("h2", parent=base, fontName="Helvetica-Bold", fontSize=10, leading=13.5,
                 alignment=rl["TA_LEFT"], spaceBefore=8, spaceAfter=4),
        "body": base,
        "small": ps("small", parent=base, fontSize=8.0, leading=11, textColor=slate),
        "bullet": ps("bullet", parent=base, leftIndent=12, bulletIndent=3, spaceAfter=3),
        "cell": ps("cell", parent=base, fontSize=8.2, leading=10.8, alignment=rl["TA_LEFT"],
                   spaceAfter=0),
        "cell_head": ps("cell_head", parent=base, fontName="Helvetica-Bold", fontSize=8.2,
                        leading=10.8, alignment=rl["TA_LEFT"], spaceAfter=0,
                        textColor=colors.white),
        "mono": ps("mono", parent=base, fontName="Courier", fontSize=8.4, leading=11.4,
                   alignment=rl["TA_LEFT"]),
        "quote": ps("quote", parent=base, leftIndent=14, rightIndent=10, fontName="Helvetica-Oblique",
                    textColor=slate, spaceBefore=4, spaceAfter=6),
    }


class _Story:
    """A tiny flowable builder, so each document function reads like the document it produces."""

    def __init__(self, rl: dict[str, Any]) -> None:
        self._rl = rl
        self._s = _styles(rl)
        self.items: list[Any] = []

    # -- text ----------------------------------------------------------------------
    def title(self, text: str, subtitle: str = "") -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["title"]))
        if subtitle:
            self.items.append(self._rl["Paragraph"](_esc(subtitle), self._s["subtitle"]))
        self.items.append(self._rl["HRFlowable"](width="100%", thickness=1.1,
                                                 color=self._rl["colors"].HexColor("#111111"),
                                                 spaceBefore=2, spaceAfter=8))
        return self

    def h1(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["h1"]))
        return self

    def h2(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["h2"]))
        return self

    def p(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["body"]))
        return self

    def small(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["small"]))
        return self

    def quote(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["quote"]))
        return self

    def mono(self, text: str) -> _Story:
        self.items.append(self._rl["Paragraph"](_esc(text), self._s["mono"]))
        return self

    def bullets(self, lines: Sequence[str], *, bullet: str = "•") -> _Story:
        for line in lines:
            self.items.append(
                self._rl["Paragraph"](_esc(line), self._s["bullet"], bulletText=bullet)
            )
        return self

    def numbered(self, lines: Sequence[str], *, start: int = 1) -> _Story:
        for offset, line in enumerate(lines):
            self.items.append(
                self._rl["Paragraph"](_esc(line), self._s["bullet"],
                                      bulletText=f"{start + offset}.")
            )
        return self

    # -- structure -----------------------------------------------------------------
    def spacer(self, height_mm: float = 3.0) -> _Story:
        self.items.append(self._rl["Spacer"](1, height_mm * self._rl["mm"]))
        return self

    def rule(self) -> _Story:
        self.items.append(self._rl["HRFlowable"](width="100%", thickness=0.6,
                                                 color=self._rl["colors"].HexColor("#9aa3ad"),
                                                 spaceBefore=4, spaceAfter=6))
        return self

    def page_break(self) -> _Story:
        self.items.append(self._rl["PageBreak"]())
        return self

    # -- tables --------------------------------------------------------------------
    def table(
        self,
        headers: Sequence[str],
        rows: Sequence[Sequence[str]],
        *,
        widths_mm: Sequence[float],
        align_right: Sequence[int] = (),
    ) -> _Story:
        """A gridded data table with a dark header row and zebra striping."""
        colors = self._rl["colors"]
        head_cells = [self._rl["Paragraph"](_esc(str(h)), self._s["cell_head"]) for h in headers]
        body_cells = [
            [self._rl["Paragraph"](_esc(str(cell)), self._s["cell"]) for cell in row]
            for row in rows
        ]
        data = [head_cells, *body_cells]
        table = self._rl["Table"](
            data,
            colWidths=[w * self._rl["mm"] for w in widths_mm],
            repeatRows=1,
            hAlign="LEFT",
        )
        style: list[tuple[Any, ...]] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#26303a")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#8e97a1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for index in range(1, len(data)):
            if index % 2 == 0:
                style.append(("BACKGROUND", (0, index), (-1, index), colors.HexColor("#f1f3f5")))
        for column in align_right:
            style.append(("ALIGN", (column, 1), (column, -1), "RIGHT"))
        table.setStyle(self._rl["TableStyle"](style))
        self.items.append(table)
        self.spacer(2.5)
        return self

    def fields(self, pairs: Sequence[tuple[str, str]], *, columns: int = 2) -> _Story:
        """A form-style block of ``label: value`` fields, laid out in ``columns`` pairs per row."""
        colors = self._rl["colors"]
        cells: list[list[Any]] = []
        row: list[Any] = []
        for label, value in pairs:
            row.append(self._rl["Paragraph"](f"<b>{_esc(label)}</b>", self._s["cell"]))
            row.append(self._rl["Paragraph"](_esc(value), self._s["cell"]))
            if len(row) >= columns * 2:
                cells.append(row)
                row = []
        if row:
            while len(row) < columns * 2:
                row.append("")
            cells.append(row)

        label_w = 28.0
        value_w = (_CONTENT_WIDTH_MM - columns * label_w) / columns
        widths = []
        for _ in range(columns):
            widths.extend([label_w, value_w])
        table = self._rl["Table"](cells, colWidths=[w * self._rl["mm"] for w in widths],
                                  hAlign="LEFT")
        table.setStyle(self._rl["TableStyle"]([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#8e97a1")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef1f4")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#eef1f4")) if columns > 1
            else ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        self.items.append(table)
        self.spacer(3)
        return self

    def image(self, png_bytes: bytes, *, width_mm: float, caption: str = "") -> _Story:
        """Place a raster image, scaled to ``width_mm`` and keeping its aspect ratio."""
        from PIL import Image as PILImage

        with PILImage.open(io.BytesIO(png_bytes)) as probe:
            width_px, height_px = probe.size
        mm = self._rl["mm"]
        height_mm = width_mm * (height_px / max(1, width_px))
        flowable = self._rl["Image"](io.BytesIO(png_bytes), width=width_mm * mm,
                                     height=height_mm * mm)
        flowable.hAlign = "LEFT"
        block: list[Any] = [flowable]
        if caption:
            block.append(self._rl["Paragraph"](_esc(caption), self._s["small"]))
        self.items.append(self._rl["KeepTogether"](block))
        self.spacer(3)
        return self


def _esc(text: str) -> str:
    """Escape the three characters ReportLab's mini-HTML parser treats as markup."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _canvas_class(rl: dict[str, Any], furniture: PageFurniture) -> type:
    """Build a canvas class that stamps ``furniture`` and a ``Page n of m`` footer on every page.

    The total page count is only known once the whole story is laid out, so pages are buffered and
    the furniture is drawn during ``save()`` — the standard ReportLab two-pass trick.
    """
    canvas_module = rl["canvas"]
    colors = rl["colors"]
    mm = rl["mm"]

    class _NumberedCanvas(canvas_module.Canvas):  # type: ignore[misc, name-defined]
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, **kwargs)
            self._pages: list[dict[str, Any]] = []

        def showPage(self) -> None:  # noqa: N802 - ReportLab API
            self._pages.append(dict(self.__dict__))
            self._startPage()

        def save(self) -> None:
            total = len(self._pages)
            for state in self._pages:
                self.__dict__.update(state)
                self._furniture(total)
                super().showPage()
            self._pages = []
            super().save()

        def _furniture(self, total: int) -> None:
            width, height = self._pagesize
            ink = colors.HexColor("#111111")
            slate = colors.HexColor("#5b636c")

            self.saveState()
            self.setFont("Helvetica-Bold", 7.6)
            self.setFillColor(ink)
            self.drawString(_MARGIN_MM * mm, height - 11 * mm, furniture.organisation.upper())
            self.setFont("Helvetica", 7.2)
            self.setFillColor(slate)
            self.drawRightString(width - _MARGIN_MM * mm, height - 11 * mm,
                                 f"{furniture.document_number}   |   "
                                 f"{furniture.document_date.isoformat()}")
            self.setStrokeColor(colors.HexColor("#9aa3ad"))
            self.setLineWidth(0.5)
            self.line(_MARGIN_MM * mm, height - 13 * mm, width - _MARGIN_MM * mm, height - 13 * mm)

            self.line(_MARGIN_MM * mm, 13 * mm, width - _MARGIN_MM * mm, 13 * mm)
            self.setFont("Helvetica", 6.8)
            self.drawString(_MARGIN_MM * mm, 9.5 * mm, furniture.classification)
            self.drawCentredString(width / 2.0, 9.5 * mm, _clip(furniture.title, 78))
            self.drawRightString(width - _MARGIN_MM * mm, 9.5 * mm,
                                 f"Page {self._pageNumber} of {total}")
            self.restoreState()

    return _NumberedCanvas


def _clip(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _render_pdf(path: Path, furniture: PageFurniture, story: _Story, rl: dict[str, Any]) -> int:
    """Write ``story`` to ``path`` and return the page count.

    Raises:
        BlobStoreError: If the PDF cannot be written.
    """
    mm = rl["mm"]
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        doc = rl["SimpleDocTemplate"](
            str(path),
            pagesize=rl["A4"],
            leftMargin=_MARGIN_MM * mm,
            rightMargin=_MARGIN_MM * mm,
            topMargin=(_MARGIN_MM + 6) * mm,
            bottomMargin=(_MARGIN_MM - 2) * mm,
            title=furniture.title,
            author=furniture.organisation,
            subject=furniture.document_number,
            creator="INDRA demo corpus generator",
            invariant=1,
        )
        doc.build(list(story.items), canvasmaker=_canvas_class(rl, furniture))
        return int(getattr(doc, "page", 0) or 0)
    except (OSError, ValueError) as exc:
        raise BlobStoreError(
            f"Could not write the generated PDF to {path}. Check that the directory exists, is "
            "writable, and that the file is not open in a PDF viewer.",
            context={"path": str(path)},
            cause=exc,
        ) from exc


# ======================================================================================
# Documents
# ======================================================================================


def _oem_manual(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """Sulzer CP 150-400 IOM extract — the authority for every threshold in the corpus."""
    p101 = facts.P101
    furniture = PageFurniture(
        organisation="Sulzer Pumps India Pvt. Ltd.",
        title=f"{p101.model} Installation, Operation and Maintenance Manual",
        document_number="IOM-CP150-400-EN-Rev4",
        document_date=date(2016, 1, 15),
        classification="OEM DOCUMENTATION — SUPPLIED WITH EQUIPMENT",
    )
    s = _Story(rl)
    s.title(
        f"{p101.manufacturer} {p101.model} — Installation, Operation and Maintenance Manual",
        f"Multistage horizontal centrifugal pump  |  Tag {p101.tag}  |  "
        f"Serial {p101.serial_number}  |  Supplied to {facts.PLANT_NAME}, {facts.PLANT_UNIT}",
    )

    s.h1("1.  Scope and application")
    s.p(
        f"This manual covers the {p101.manufacturer} {p101.model} multistage boiler feed water pump "
        f"supplied under order BVPL/2015/UT-2/091 and installed as tag {p101.tag} in the "
        f"{p101.unit} of {facts.PLANT_NAME}. A second identical machine is installed as "
        f"{facts.P102.tag} in duty-standby configuration. The instructions in Section 7 are "
        "mandatory: warranty and the stated bearing life are conditional on the replacement "
        "criteria of Section 7.4 being observed."
    )
    s.p(
        f"The pump takes suction from the deaerator storage vessel {facts.V201.tag} and discharges "
        f"through the feed water pre-heater {facts.E301.tag} to boiler B-401. Forced-feed "
        f"lubrication is supplied by the auxiliary lube oil pump {facts.LP101A.tag} mounted on the "
        "same skid. Refer to P&ID PID-U2-1010 Rev 3 for the process connections."
    )

    s.h1("2.  Technical data")
    s.table(
        ["Parameter", "Value"],
        [[key, value] for key, value in p101.specifications.items()],
        widths_mm=[58, _CONTENT_WIDTH_MM - 58],
    )

    s.h1("3.  Installation")
    s.bullets((
        "Grout the baseplate and confirm flatness within 0.05 mm/m before coupling alignment.",
        "Align the pump to the driver within 0.05 mm rim and 0.03 mm face, cold, with piping "
        "connected. Record the readings in the commissioning file.",
        "Suction piping from V-201 must fall continuously to the pump suction flange. A pocket in "
        "the suction line will cause vapour binding at low deaerator level.",
        "Do not use the casing to support the discharge pipework. Flange loads must not exceed "
        "the values in Appendix C.",
    ))

    s.h1("4.  Operation")
    s.p(
        "Start against a closed discharge valve with the minimum-flow recirculation line open. "
        "Continuous operation below 40 m3/h will overheat the casing and shorten bearing life "
        "through increased radial thrust. Auto changeover to the standby machine is initiated on "
        "low discharge pressure at PI-1015."
    )

    s.page_break()
    s.h1("5.  Lubrication system")
    s.p(
        f"Both bearing housings are lubricated by forced feed from {facts.LP101A.tag} "
        f"({facts.LP101A.manufacturer} {facts.LP101A.model}), rated 18 l/min at 2.1 bar with "
        "ISO VG 46 turbine oil. The lube oil header pressure is indicated at PI-1016."
    )
    s.table(
        ["Lubrication parameter", "Normal", "Alarm", "Trip"],
        [
            ["Header pressure (bar)", "2.1", f"{facts.OEM_LUBE_PRESSURE_MIN_BAR} low", "1.1 low-low"],
            ["Oil temperature (degC)", "45 - 60", "70", "80"],
            ["Oil cleanliness (ISO 4406)", "16/14/11", "18/16/13", "20/18/15"],
            ["Iron content (ppm)", "below 10", "20", "35"],
        ],
        widths_mm=[62, 34, 40, _CONTENT_WIDTH_MM - 136],
        align_right=(1, 2, 3),
    )
    s.p(
        f"Loss of lube oil header pressure below {facts.OEM_LUBE_PRESSURE_MIN_BAR} bar removes the "
        "oil film from the rolling elements within minutes at rated speed. Sustained operation "
        "below this value will destroy the non-drive-end bearing."
    )

    s.h1("6.  Condition monitoring")
    s.table(
        ["Measured parameter", "Location", "Alarm", "Trip", "Instrument"],
        [
            ["Overall vibration (mm/s RMS)", "NDE bearing housing",
             f"{facts.OEM_VIBRATION_LIMIT_MM_S}", "9.5", "VT-1011"],
            ["Bearing metal temperature (degC)", "NDE and DE housings",
             "85", f"{facts.OEM_BEARING_TEMP_LIMIT_C:.0f}", "TE-1011A/B"],
            ["Discharge pressure (barg)", "Pump discharge", "38 low", "35 low-low", "PI-1015"],
            ["Lube oil header pressure (bar)", "Bearing oil header",
             f"{facts.OEM_LUBE_PRESSURE_MIN_BAR} low", "1.1 low-low", "PI-1016"],
        ],
        widths_mm=[46, 40, 26, 26, _CONTENT_WIDTH_MM - 138],
        align_right=(2, 3),
    )
    s.p(
        "Vibration spectra shall be recorded quarterly with a data collector capable of resolving "
        "1x and 2x running speed. A rising 1x component with a 2x sideband at the non-drive end is "
        "the characteristic signature of outer-race wear on the angular-contact bearing and "
        "precedes measurable temperature rise by several weeks."
    )

    s.page_break()
    s.h1("7.  Maintenance")
    s.h2("7.1  Routine schedule")
    s.table(
        ["Task", "Interval", "Reference"],
        [
            ["Lube oil sample and analysis", "Quarterly", "Section 5"],
            ["Vibration survey with spectrum", "Quarterly", "Section 6"],
            ["Bearing wear assessment", "Quarterly", "Section 7.4"],
            ["Mechanical seal inspection", "Annually", "Section 7.6"],
            ["Full overhaul", "5 years or on condition", "Section 8"],
        ],
        widths_mm=[62, 40, _CONTENT_WIDTH_MM - 102],
    )

    s.h2("7.4  Bearing replacement criteria — MANDATORY")
    s.p(
        f"Bearing replacement is mandatory once measured wear reaches "
        f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% of the permissible limit defined in Table 7-2. "
        f"Wear at or above {facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% shall not be carried into the "
        "next operating quarter under any circumstances, and shall not be deferred against "
        "production requirements. Replacement is to be executed by a competent person following "
        "the plant's own bearing replacement procedure."
    )
    s.table(
        ["Wear (% of limit)", "Classification", "Required action"],
        [
            ["Below 60", "Normal", "Continue quarterly monitoring."],
            ["60 to 74", "Elevated", "Increase monitoring to monthly. Plot the trend against the "
                                     f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% limit."],
            ["75 to 84", "Pre-replacement", "Raise the replacement work order and order the "
                                            "bearing set. Confirm the spare is on site."],
            [f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f} and above", "Replacement mandatory",
             "Replace at the next available opportunity. Do not carry into the next quarter."],
        ],
        widths_mm=[30, 34, _CONTENT_WIDTH_MM - 64],
    )
    s.p(
        "Wear is assessed by comparing outer-race shell thickness and load-zone scoring against "
        "the reference photographs in Appendix D. Where the wear pattern matches the scoring "
        "signature of a previous seizure on the same machine class, treat the reading as "
        "pre-replacement irrespective of the measured percentage."
    )

    s.h2("7.6  Bearing arrangement")
    s.bullets((
        f"Drive end: {p101.specifications.get('DE bearing', 'SKF 6316 C3')} — deep-groove ball "
        "bearing, radial load only.",
        f"Non-drive end: {p101.specifications.get('NDE bearing', 'SKF 7314 BECBM')} — angular "
        "contact, carries the residual axial thrust. This is the bearing that fails first.",
        "Never flame-heat a bearing. Induction heating to a maximum of 110 degC only.",
    ))

    s.h1("8.  Spare parts and process interfaces")
    s.table(
        ["Item", "Part number", "Recommended stock"],
        [
            ["NDE bearing (angular contact)", "SKF 7314 BECBM", "1 set"],
            ["DE bearing (deep groove)", "SKF 6316 C3", "1 set"],
            ["Cartridge mechanical seal", "SLZ-MS-150-400", "1"],
            ["Bearing housing lip seals", "SLZ-LS-314", "2 sets"],
            ["Lube oil pump rotor/idler set", "RD-RGP-18-ROT", "1"],
        ],
        widths_mm=[62, 46, _CONTENT_WIDTH_MM - 108],
    )
    s.small(
        f"Process interfaces: suction from {facts.V201.tag} ({facts.V201.name}); discharge to "
        f"{facts.E301.tag} ({facts.E301.name}); lubrication from {facts.LP101A.tag}. "
        "Refer to P&ID PID-U2-1010 Rev 3."
    )
    return furniture, s


def _work_order(rl: dict[str, Any], *, annotation_png: bytes) -> tuple[PageFurniture, _Story]:
    """WO-2024-0342 — the open investigation the demo's first diagnostic question lands on."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"Work Order {facts.WORK_ORDER_ID} — P-101 Vibration Anomaly Investigation",
        document_number=facts.WORK_ORDER_ID,
        document_date=facts.WORK_ORDER_DATE,
        classification="MAINTENANCE RECORD — CMMS EXPORT",
    )
    s = _Story(rl)
    s.title(
        f"Maintenance Work Order {facts.WORK_ORDER_ID}",
        f"{facts.PLANT_UNIT}  |  Mechanical Maintenance  |  Raised "
        f"{facts.WORK_ORDER_DATE.isoformat()}",
    )
    s.fields((
        ("Work order", facts.WORK_ORDER_ID),
        ("Status", facts.WORK_ORDER_STATUS),
        ("Equipment", f"{facts.P101.tag} — {facts.P101.name}"),
        ("Priority", facts.WORK_ORDER_PRIORITY),
        ("Location", facts.P101.location),
        ("Work type", "Investigation (no corrective work authorised)"),
        ("Assigned to", f"{facts.RAJESH_KUMAR.name} ({facts.RAJESH_KUMAR.employee_id})"),
        ("Raised by", f"{facts.ANIL_DESHMUKH.name} ({facts.ANIL_DESHMUKH.employee_id})"),
        ("Trade", "Rotating equipment"),
        ("Planned hours", "4.0"),
        ("Related documents", f"{facts.INSPECTION_ID}, ShiftLog {facts.SHIFT_LOG_DATE.isoformat()}"),
        ("Cost centre", "UT2-MECH-01"),
    ))

    s.h1("1.  Problem description")
    s.p(
        f"Vibration anomaly reported on {facts.P101.tag} following two high-vibration alarms on "
        f"the night shift of {facts.SHIFT_LOG_DATE.isoformat()}. The B Shift supervisor requested "
        "an investigation of the non-drive-end bearing condition, the lube oil supply and the "
        "discharge pressure trend. This work order authorises investigation and reporting only."
    )

    s.h1("2.  Findings")
    s.numbered(facts.WORK_ORDER_FINDINGS)

    s.h1("3.  Field measurements")
    s.table(
        ["Parameter", "Measured", "OEM limit", "Percent of limit", "Instrument / method"],
        [
            ["NDE bearing wear", f"{facts.MEASURED_BEARING_WEAR_PCT:.0f}%",
             f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}%",
             f"{100 * facts.MEASURED_BEARING_WEAR_PCT / facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}%",
             "Shell-thickness comparison, Appendix D reference set"],
            ["Overall vibration", "6.9 mm/s RMS", f"{facts.OEM_VIBRATION_LIMIT_MM_S} mm/s",
             f"{100 * 6.9 / facts.OEM_VIBRATION_LIMIT_MM_S:.0f}%", "VT-1011, portable collector"],
            ["NDE bearing housing temp.", "81 degC",
             f"{facts.OEM_BEARING_TEMP_LIMIT_C:.0f} degC",
             f"{100 * 81 / facts.OEM_BEARING_TEMP_LIMIT_C:.0f}%", "Contact pyrometer"],
            ["Lube oil header pressure", "1.8 bar",
             f"{facts.OEM_LUBE_PRESSURE_MIN_BAR} bar min", "129% of minimum", "PI-1016"],
        ],
        widths_mm=[36, 26, 26, 26, _CONTENT_WIDTH_MM - 114],
        align_right=(1, 2, 3),
    )

    s.spacer(2)
    s.h2("Technician's margin note (scanned from the field copy)")
    s.image(annotation_png, width_mm=76,
            caption="Handwritten annotation on the field copy of this work order, initialled RK. "
                    "Transcribed value carried into Section 3 above.")

    s.page_break()
    s.h1("4.  Scheduling status")
    s.p(facts.WORK_ORDER_NO_REPLACEMENT_SENTENCE)
    s.table(
        ["Follow-on work", "Status", "Reference"],
        [
            ["NDE / DE bearing replacement", "NOT RAISED — not scheduled", facts.SOP_ID],
            ["Lube oil sample and analysis", "Requested, not yet sampled", "Lab request LR-2024-0611"],
            ["VT-1011 alarm bypass removal", "Open with instrumentation", "Shift log "
                                                                          f"{facts.SHIFT_LOG_DATE.isoformat()}"],
            ["LP-101A discharge trend review", "Open with reliability", facts.RCA_ID],
        ],
        widths_mm=[54, 52, _CONTENT_WIDTH_MM - 106],
    )

    s.h1("5.  Recommendations")
    s.numbered(facts.WORK_ORDER_RECOMMENDATIONS)

    s.h1("6.  History considered")
    s.p(
        f"The {facts.INSPECTION_DATE.isoformat()} quarterly inspection ({facts.INSPECTION_ID}) "
        f"recorded 71% wear and noted a wear pattern similar to the 2022 failure. The 2022 event "
        f"({facts.INCIDENT_ID}) was an NDE bearing seizure that cost "
        f"{facts.INCIDENT_DOWNTIME_HOURS:.0f} hours of lost steam raising and "
        f"{facts.INCIDENT_COST_TEXT}. Root cause analysis {facts.RCA_ID} attributed it to "
        "lubrication failure originating in LP-101A."
    )

    s.h1("7.  Sign-off")
    s.fields((
        ("Executed by", f"{facts.RAJESH_KUMAR.name}, {facts.RAJESH_KUMAR.role}"),
        ("Date", facts.WORK_ORDER_DATE.isoformat()),
        ("Reviewed by", f"{facts.D_KRISHNAN.name}, {facts.D_KRISHNAN.role}"),
        ("Status at close", f"{facts.WORK_ORDER_STATUS} — investigation only"),
    ))
    s.small(
        "This work order remains OPEN. No corrective maintenance has been authorised against it. "
        "Any bearing replacement requires a separate work order raised against "
        f"{facts.SOP_ID} {facts.SOP_REVISION}."
    )
    return furniture, s


def _inspection(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """The quarterly inspection whose one sentence bridges 2024 condition to the 2022 failure."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"Quarterly Mechanical Inspection {facts.INSPECTION_ID} — {facts.P101.tag}",
        document_number=facts.INSPECTION_ID,
        document_date=facts.INSPECTION_DATE,
        classification="INSPECTION RECORD — ASSET RELIABILITY",
    )
    s = _Story(rl)
    s.title(
        f"Quarterly Mechanical Inspection Report {facts.INSPECTION_ID}",
        f"{facts.P101.tag} — {facts.P101.name}  |  Inspected "
        f"{facts.INSPECTION_DATE.isoformat()} by {facts.PRIYA_SHARMA.name}",
    )
    s.fields((
        ("Report number", facts.INSPECTION_ID),
        ("Inspection date", facts.INSPECTION_DATE.isoformat()),
        ("Equipment", f"{facts.P101.tag} ({facts.P101.manufacturer} {facts.P101.model})"),
        ("Criticality", f"Class {facts.P101.criticality}"),
        ("Inspector", f"{facts.PRIYA_SHARMA.name}, {facts.PRIYA_SHARMA.role}"),
        ("Witness", f"{facts.RAJESH_KUMAR.name}"),
        ("Running hours", "63,480 h since 2016 commissioning"),
        ("Next inspection due", "2024-06-14"),
    ))

    s.h1("1.  Scope and method")
    s.p(
        "Quarterly condition assessment of the non-drive-end and drive-end bearings, lubrication "
        "system, mechanical seal and coupling alignment, carried out with the machine running at "
        "steady duty and again during the 90-minute shutdown window granted by operations. Wear "
        "assessment by outer-race shell-thickness comparison against the OEM Appendix D reference "
        "set; vibration by portable collector with 1x/2x spectral capture; oil condition by "
        "laboratory analysis of a sample drawn from the bearing return line."
    )

    s.h1("2.  Findings")
    s.numbered(facts.INSPECTION_FINDINGS)

    s.h1("3.  Measurements against limits")
    s.table(
        ["Parameter", "This inspection", "Previous (2023-09-19)", "OEM limit", "Assessment"],
        [
            ["NDE bearing wear (%)", "71", "52",
             f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}", "Elevated — trend accelerating"],
            ["Overall vibration (mm/s RMS)", "5.8", "4.4",
             f"{facts.OEM_VIBRATION_LIMIT_MM_S}", "Within alarm, rising"],
            ["NDE housing temperature (degC)", "74", "68",
             f"{facts.OEM_BEARING_TEMP_LIMIT_C:.0f}", "Within limit"],
            ["Lube oil header pressure (bar)", "1.9", "2.0",
             f"{facts.OEM_LUBE_PRESSURE_MIN_BAR} min", "Within limit, drifting down"],
            ["Oil iron content (ppm)", "22", "9", "20 alarm", "Above alarm"],
        ],
        widths_mm=[42, 26, 32, 24, _CONTENT_WIDTH_MM - 124],
        align_right=(1, 2, 3),
    )

    s.page_break()
    s.h1("4.  Wear pattern assessment")
    s.p(facts.INSPECTION_SIMILARITY_NOTE)
    s.quote(
        "The scoring on the outer race load zone and the 1x/2x spectral signature are the same "
        "pattern recorded before the August 2022 seizure of this machine. On the OEM's own "
        "criteria (Section 7.4) a wear pattern matching a previous seizure signature is to be "
        "treated as pre-replacement irrespective of the measured percentage."
    )

    s.h1("5.  Bearing wear trend since the 2022 replacement")
    wear = facts.readings_for(facts.P101.tag, "bearing_wear_pct")
    s.table(
        ["Date", "Wear (%)", "Percent of OEM limit", "Source"],
        [
            [r.measured_on.isoformat(), f"{r.value:.0f}",
             f"{100 * r.value / facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}%", r.source_document]
            for r in wear if r.measured_on <= facts.INSPECTION_DATE
        ],
        widths_mm=[28, 24, 40, _CONTENT_WIDTH_MM - 92],
        align_right=(1, 2),
    )
    s.p(
        "Wear has advanced roughly 7 percentage points per quarter since the bearing set was "
        f"renewed on 2022-08-26. On that gradient the {facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% "
        "replacement limit is reached in September 2024."
    )

    s.h1("6.  Conclusion and recommendation")
    s.bullets((
        "The machine is fit for continued service to the next quarterly inspection, conditional on "
        "the recommendations below being actioned.",
        "Plan the bearing replacement now rather than on reaching the limit: the spare set has a "
        "six-week lead time and the last replacement took a 240-minute outage window.",
        "Restore quarterly lube oil analysis and investigate the iron content rise from 9 ppm to "
        "22 ppm, which on this machine previously indicated auxiliary lube pump wear.",
        "Plot bearing wear against the OEM limit on every subsequent report — corrective action 5 "
        f"of {facts.RCA_ID} requires it.",
    ))
    s.small(
        f"Distribution: {facts.D_KRISHNAN.name} (Maintenance Manager), {facts.RAJESH_KUMAR.name} "
        f"(Senior Technician), {facts.S_RAMASWAMY.name} (Operations Foreman). Filed against "
        f"{facts.P101.tag} in the asset register."
    )
    return furniture, s


def _shift_log(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """The night shift log: the operational signal nobody would connect to a bearing."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"Shift Log — {facts.SHIFT_LOG_SHIFT}, {facts.SHIFT_LOG_DATE.isoformat()}",
        document_number=f"SL-{facts.SHIFT_LOG_DATE.strftime('%Y%m%d')}-B",
        document_date=facts.SHIFT_LOG_DATE,
        classification="OPERATIONS RECORD — CONTROL ROOM",
    )
    s = _Story(rl)
    s.title(
        f"Control Room Shift Log — {facts.PLANT_UNIT}",
        f"{facts.SHIFT_LOG_SHIFT}  |  {facts.SHIFT_LOG_DATE.isoformat()}",
    )
    s.fields((
        ("Shift", facts.SHIFT_LOG_SHIFT),
        ("Date", facts.SHIFT_LOG_DATE.isoformat()),
        ("Supervisor", f"{facts.ANIL_DESHMUKH.name} ({facts.ANIL_DESHMUKH.employee_id})"),
        ("Panel operator", f"{facts.SURESH_PATIL.name} ({facts.SURESH_PATIL.employee_id})"),
        ("Unit status", "Unit-2 steam raising normal throughout"),
        ("Handover to", "C Shift, 06:00"),
    ))

    s.h1("1.  Log entries")
    s.table(
        ["Time", "Entry"],
        [[time, text] for time, text in facts.SHIFT_LOG_ENTRIES],
        widths_mm=[18, _CONTENT_WIDTH_MM - 18],
    )

    s.h1("2.  Alarm summary")
    s.table(
        ["Time", "Tag", "Alarm", "Value", "Operator action"],
        [
            ["02:14", "VT-1011", "P-101 high vibration", "6.9 mm/s",
             "Alarm bypassed to stop repeat annunciation. No work order raised."],
            ["03:47", "VT-1011", "P-101 high vibration", "6.9 mm/s",
             "Alarm bypassed a second time. Supervisor informed at 04:10."],
            ["01:20", "PI-1015", "P-101 discharge pressure high deviation", "42.1 barg",
             "Monitored. E-301 tube-side fouling suspected."],
        ],
        widths_mm=[16, 20, 44, 22, _CONTENT_WIDTH_MM - 102],
    )
    s.p(
        f"{facts.ALARM_BYPASS_NOTE}. Both bypasses were applied at the DCS operator station "
        "without a maintenance notification being raised at the time. The bypasses remained in "
        "force at handover."
    )

    s.h1("3.  Open items carried to C Shift")
    s.bullets((
        "P-101 vibration elevated and VT-1011 alarm bypassed twice — raise a maintenance "
        "notification at day-shift handover.",
        "P-101 NDE bearing housing temperature 81 degC, up 7 degC across the shift.",
        "E-301 tube-side fouling suspected; discharge pressure 0.6 barg above the shift average.",
        "V-201 deaerator level and dissolved oxygen normal; no action.",
    ))

    s.h1("4.  Shift signatures")
    s.fields((
        ("Outgoing supervisor", f"{facts.ANIL_DESHMUKH.name}"),
        ("Time", "06:00"),
        ("Panel operator", f"{facts.SURESH_PATIL.name}"),
        ("Time", "06:00"),
    ))
    s.small(
        "Control room logs are retained for three years under the plant's records procedure and "
        "are admissible evidence in incident investigation."
    )
    return furniture, s


def _incident(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """The 2022 seizure. The business case for acting on the 2024 pattern."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"Incident Report {facts.INCIDENT_ID} — P-101 NDE Bearing Seizure",
        document_number=facts.INCIDENT_ID,
        document_date=facts.INCIDENT_DATE,
        classification="INCIDENT RECORD — RESTRICTED",
    )
    s = _Story(rl)
    s.title(
        f"Plant Incident Report {facts.INCIDENT_ID}",
        f"{facts.P101.tag} non-drive-end bearing seizure  |  {facts.INCIDENT_DATE.isoformat()}  |  "
        f"{facts.PLANT_UNIT}",
    )
    s.fields((
        ("Incident number", facts.INCIDENT_ID),
        ("Date of event", facts.INCIDENT_DATE.isoformat()),
        ("Equipment", f"{facts.P101.tag} — {facts.P101.name}"),
        ("Failure mode", facts.INCIDENT_FAILURE_MODE),
        ("Classification", "Category 2 — significant production loss, no injury"),
        ("Reportable", "No injury; not reportable under Section 88"),
        ("Downtime", f"{facts.INCIDENT_DOWNTIME_HOURS:.0f} hours"),
        ("Total cost", facts.INCIDENT_COST_TEXT),
        ("Investigation lead", f"{facts.PRIYA_SHARMA.name}"),
        ("RCA reference", facts.RCA_ID),
    ))

    s.h1("1.  Summary")
    s.p(
        f"At 04:52 on {facts.INCIDENT_DATE.isoformat()} the non-drive-end bearing of "
        f"{facts.P101.tag} seized while the machine was running at rated duty. The shaft locked, "
        "the motor tripped on overcurrent, and Unit-2 lost boiler feed water supply. The standby "
        f"machine {facts.P102.tag} was started manually at 05:04 but tripped on high vibration; "
        "steam raising was lost for a total of "
        f"{facts.INCIDENT_DOWNTIME_HOURS:.0f} hours until {facts.P101.tag} was returned to service "
        "with a new bearing set."
    )

    s.h1("2.  Timeline")
    s.table(
        ["Date / time", "Event"],
        [
            ["2022-08-14", "NDE bearing temperature 68 degC at routine round. Recorded, no action."],
            ["2022-08-19 22:40", "VT-1011 high vibration alarm (8.1 mm/s). Acknowledged on the "
                                 "night shift. No work order raised."],
            ["2022-08-19 23:55", "VT-1011 high vibration alarm again (8.9 mm/s). Acknowledged. No "
                                 "corrective action."],
            ["2022-08-20 02:30", "Lube oil header pressure 1.2 bar, below the 1.4 bar low-low "
                                 "limit. No alarm configured on PI-1016 at the time."],
            ["2022-08-20 04:52", "NDE bearing seized. Shaft locked. Motor tripped on overcurrent."],
            ["2022-08-20 05:04", "P-102 started manually; tripped on high vibration after 6 "
                                 "minutes."],
            ["2022-08-20 05:20", "Unit-2 steam raising lost. Deaerator V-201 level held on "
                                 "condensate return."],
            ["2022-08-20 18:52", "P-101 returned to service after bearing replacement. Total "
                                 f"downtime {facts.INCIDENT_DOWNTIME_HOURS:.0f} hours."],
        ],
        widths_mm=[32, _CONTENT_WIDTH_MM - 32],
    )

    s.page_break()
    s.h1("3.  Precursors identified in the investigation")
    s.numbered(facts.INCIDENT_PRECURSORS)
    s.p(
        "Each precursor was individually visible in plant records before the failure. None was "
        "individually alarming. The combination was never assembled by any single person or "
        "system, which is the central finding of this report."
    )

    s.h1("4.  Consequences and cost")
    s.table(
        ["Cost element", "Amount (Rs)", "Basis"],
        [
            ["Lost steam production", "14,20,000",
             f"{facts.INCIDENT_DOWNTIME_HOURS:.0f} h at the Unit-2 contribution rate"],
            ["Emergency bearing set and freight", "4,80,000", "Air freight from the OEM"],
            ["Contract labour, 3 shifts", "2,60,000", "Rotating equipment crew"],
            ["Shaft skim and re-metalling", "2,15,000", "External workshop"],
            ["Lube oil flush and refill", "1,25,000", "ISO VG 46, 220 l"],
            ["Total", "25,00,000", facts.INCIDENT_COST_TEXT],
        ],
        widths_mm=[62, 34, _CONTENT_WIDTH_MM - 96],
        align_right=(1,),
    )
    s.p(
        f"Total cost {facts.INCIDENT_COST_TEXT} against {facts.INCIDENT_DOWNTIME_HOURS:.0f} hours "
        "of downtime. No injury occurred. There was no loss of containment and no environmental "
        "consequence."
    )

    s.h1("5.  Immediate actions taken")
    s.bullets((
        f"{facts.P101.tag} isolated and the NDE and DE bearing sets replaced per {facts.SOP_ID} by "
        f"{facts.RAJESH_KUMAR.name}.",
        f"{facts.LP101A.tag} auxiliary lube oil pump found with heavy internal wear and overhauled "
        "on 2022-08-24; discharge restored to 2.1 bar.",
        f"{facts.P102.tag} standby machine inspected; alignment corrected and returned to standby "
        "duty.",
        "Lube oil system flushed and recharged; oil sample sent for analysis.",
        f"Root cause analysis {facts.RCA_ID} commissioned; issued {facts.RCA_DATE.isoformat()}.",
    ))

    s.h1("6.  Equipment affected")
    s.table(
        ["Tag", "Description", "Role in the event"],
        [
            [facts.P101.tag, facts.P101.name, "Failed machine — NDE bearing seizure"],
            [facts.LP101A.tag, facts.LP101A.name, "Root cause — lube oil pressure decay"],
            [facts.P102.tag, facts.P102.name, "Standby — tripped on start, extended the outage"],
            [facts.V201.tag, facts.V201.name, "Suction source — level held on condensate return"],
        ],
        widths_mm=[22, 62, _CONTENT_WIDTH_MM - 84],
    )
    s.small(
        f"Prepared by {facts.PRIYA_SHARMA.name}. Reviewed by {facts.D_KRISHNAN.name}. "
        f"Witness statements from {facts.RAJESH_KUMAR.name} and {facts.ANIL_DESHMUKH.name} are "
        "held in the investigation file."
    )
    return furniture, s


def _rca(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """The root cause analysis: names the mechanism, so the Copilot can explain *why*."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"Root Cause Analysis {facts.RCA_ID} — P-101 NDE Bearing Seizure",
        document_number=facts.RCA_ID,
        document_date=facts.RCA_DATE,
        classification="INVESTIGATION RECORD — RESTRICTED",
    )
    s = _Story(rl)
    s.title(
        f"Root Cause Analysis {facts.RCA_ID}",
        f"{facts.P101.tag} non-drive-end bearing seizure of {facts.INCIDENT_DATE.isoformat()}  |  "
        f"Issued {facts.RCA_DATE.isoformat()}",
    )
    s.fields((
        ("RCA number", facts.RCA_ID),
        ("Incident reference", facts.INCIDENT_ID),
        ("Equipment", f"{facts.P101.tag}, {facts.LP101A.tag}"),
        ("Methodology", "Five-why with evidence verification"),
        ("Team lead", f"{facts.PRIYA_SHARMA.name}, {facts.PRIYA_SHARMA.role}"),
        ("Team", f"{facts.RAJESH_KUMAR.name}, {facts.D_KRISHNAN.name}"),
        ("Issued", facts.RCA_DATE.isoformat()),
        ("Status", "Closed — corrective actions tracked to completion"),
    ))

    s.h1("1.  Root cause statement")
    s.p(facts.RCA_ROOT_CAUSE)

    s.h1("2.  Five-why chain")
    s.table(
        ["#", "Question", "Answer"],
        [[str(index + 1), question, answer]
         for index, (question, answer) in enumerate(facts.RCA_FIVE_WHYS)],
        widths_mm=[8, 52, _CONTENT_WIDTH_MM - 60],
    )

    s.h1("3.  Evidence examined")
    s.bullets((
        "Failed bearing: rolling elements welded to the outer race at the load zone; cage "
        "fragmented. Consistent with oil-film loss, not with overload or misalignment.",
        f"{facts.LP101A.tag} strip-down: rotor and idler clearances opened to 0.31 mm against a "
        "0.06 mm build figure. Volumetric efficiency calculated at approximately 55%.",
        "DCS trend: lube oil header pressure decayed from 2.1 bar to 1.2 bar over six days.",
        "Alarm history: two VT-1011 high-vibration alarms acknowledged on the night shift of "
        "19 August 2022 with no work order raised.",
        "Oil analysis: last sample November 2021, 18 months before the event, against a 3-month "
        "standard interval.",
        "June 2022 inspection record: bearing wear 84% against the "
        f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% OEM replacement limit, not escalated.",
    ))

    s.page_break()
    s.h1("4.  Contributing factors")
    s.numbered(facts.RCA_CONTRIBUTING_FACTORS)

    s.h1("5.  Corrective actions")
    s.table(
        ["#", "Action", "Owner", "Target date"],
        [[str(index + 1), action, owner, when.isoformat()]
         for index, (action, owner, when) in enumerate(facts.RCA_CORRECTIVE_ACTIONS)],
        widths_mm=[8, 92, 30, _CONTENT_WIDTH_MM - 130],
    )

    s.h1("6.  Failure mechanism, in plain language")
    s.p(
        "The auxiliary lube oil pump wore out quietly. Nothing watched it, because it had no "
        "condition-monitoring task and no low-pressure alarm. As its internal clearances opened "
        "up, the oil header pressure fell a little each week. Below 1.4 bar the oil film in the "
        "non-drive-end bearing could no longer separate the rolling elements from the race. The "
        "bearing ran metal-to-metal, heated, and welded itself solid. The vibration alarms that "
        "should have interrupted this were acknowledged twice on a night shift and no work order "
        "was raised."
    )
    s.quote(
        "The failure did not begin on 20 August 2022. It began when a 3-month oil analysis "
        "interval was allowed to run for 18 months, and it became inevitable when two alarms were "
        "acknowledged without action."
    )

    s.h1("7.  Effectiveness review")
    s.table(
        ["Action", "Completed", "Evidence", "Effective"],
        [
            ["LP-101A overhaul", "2022-08-24", "Work order MR-2022-0824", "Yes — 2.1 bar restored"],
            ["Bearing replacement", "2022-08-26", "Work order MR-2022-0826",
             "Yes — post-job vibration 3.1 mm/s"],
            ["Quarterly oil analysis reinstated", "2022-09-15", "Lab schedule LS-2022-09",
             "Partial — lapsed again in 2023"],
            ["PI-1016 low-pressure alarm added", "2022-09-30", "DCS alarm list Rev 12", "Yes"],
            ["Wear plotted against the OEM limit", "2022-10-14", "Inspection template Rev 3",
             "Partial — not applied on every report"],
        ],
        widths_mm=[46, 24, 40, _CONTENT_WIDTH_MM - 110],
    )
    s.small(
        "Two corrective actions are recorded as only partially effective. Both concern the "
        "detection of exactly the condition now present on this machine."
    )
    return furniture, s


def _retirement_email(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """The HR e-mail nobody would think to search — where the knowledge cliff hides."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title="E-mail — Superannuation schedule FY2024-27",
        document_number="HR/SUP/2024/117",
        document_date=facts.RETIREMENT_EMAIL_DATE,
        classification="HUMAN RESOURCES — INTERNAL",
    )
    s = _Story(rl)
    s.title("Internal E-mail", "Printed from the plant mail archive for the records file")
    s.fields((
        ("From", f"{facts.MEERA_IYER.name} <{facts.MEERA_IYER.contact}>"),
        ("Sent", f"{facts.RETIREMENT_EMAIL_DATE.isoformat()} 11:24 IST"),
        ("To", f"{facts.D_KRISHNAN.name} <{facts.D_KRISHNAN.contact}>"),
        ("Cc", f"{facts.PRIYA_SHARMA.name} <{facts.PRIYA_SHARMA.contact}>"),
        ("Subject", facts.RETIREMENT_EMAIL_SUBJECT),
        ("Reference", "HR/SUP/2024/117"),
    ))

    s.h1("Message")
    s.p("Dear Krishnan,")
    s.p(
        "Please find below the confirmed superannuation schedule for Mechanical Maintenance and "
        "Utilities Operations covering financial years 2024-25 to 2026-27. Departmental heads are "
        "requested to confirm handover and knowledge-capture plans for each name by the end of "
        "next month."
    )
    s.p(facts.RETIREMENT_HEADLINE)

    s.h1("Superannuation schedule")
    s.table(
        ["Employee", "ID", "Role", "Service (years)", "Superannuation date"],
        [
            [facts.RAJESH_KUMAR.name, facts.RAJESH_KUMAR.employee_id, facts.RAJESH_KUMAR.role,
             f"{facts.RAJESH_KUMAR.years_experience:.0f}", "2027-03-31 (March 2027)"],
            [facts.S_RAMASWAMY.name, facts.S_RAMASWAMY.employee_id, facts.S_RAMASWAMY.role,
             f"{facts.S_RAMASWAMY.years_experience:.0f}", "2024-11-30"],
            ["B. N. Joshi", "EMP-1042", "Boiler Attendant — A Shift", "27", "2026-07-31"],
        ],
        widths_mm=[30, 20, 52, 22, _CONTENT_WIDTH_MM - 124],
    )

    s.h1("Knowledge risk flagged by Asset Reliability")
    s.numbered(facts.RETIREMENT_KNOWLEDGE_RISK)

    s.page_break()
    s.h1("Requested actions")
    s.table(
        ["Action", "Owner", "Due"],
        [
            ["Nominate a successor for the Sulzer CP-series bearing replacement certification",
             facts.D_KRISHNAN.name, "2024-08-31"],
            ["Schedule knowledge-capture interviews for P-101, P-102 and LP-101A",
             facts.PRIYA_SHARMA.name, "2024-09-30"],
            ["Write a method statement for the P-101 bearing replacement beyond the generic SOP",
             facts.RAJESH_KUMAR.name, "2024-10-31"],
            ["Confirm V-201 statutory inspection cover after the November superannuation",
             facts.D_KRISHNAN.name, "2024-10-15"],
        ],
        widths_mm=[92, 32, _CONTENT_WIDTH_MM - 124],
    )
    s.p(
        f"{facts.RAJESH_KUMAR.name} has been with the plant since 2004 and is the named competent "
        "person on every bearing replacement executed on the Unit-2 feed water pumps. "
        f"{facts.S_RAMASWAMY.name} holds the equivalent operating knowledge for the deaerator "
        f"{facts.V201.tag} and its statutory examination history, and leaves us in November this "
        "year. I would be grateful if both handovers were treated as priority."
    )
    s.p("Regards,")
    s.p(f"{facts.MEERA_IYER.name}")
    s.small(f"{facts.MEERA_IYER.role}, {facts.PLANT_NAME}  |  {facts.MEERA_IYER.contact}")
    return furniture, s


def _sop(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """SOP-MECH-014: the procedural answer, with real steps and a real time budget."""
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"{facts.SOP_ID} {facts.SOP_REVISION} — {facts.SOP_TITLE}",
        document_number=f"{facts.SOP_ID} {facts.SOP_REVISION}",
        document_date=facts.SOP_EFFECTIVE_DATE,
        classification="CONTROLLED DOCUMENT — MECHANICAL MAINTENANCE",
    )
    hold_points = [step for step in facts.SOP_STEPS if step.hold_point]
    s = _Story(rl)
    s.title(
        f"{facts.SOP_ID} — {facts.SOP_TITLE}",
        f"{facts.SOP_REVISION}  |  Effective {facts.SOP_EFFECTIVE_DATE.isoformat()}  |  "
        f"{len(facts.SOP_STEPS)} steps  |  {facts.SOP_TOTAL_MINUTES} minutes total",
    )
    s.fields((
        ("Procedure", facts.SOP_ID),
        ("Revision", facts.SOP_REVISION),
        ("Effective date", facts.SOP_EFFECTIVE_DATE.isoformat()),
        ("Review due", "2026-02-01"),
        ("Applies to", f"{facts.P101.tag}, {facts.P102.tag}, {facts.P105.tag} "
                       f"({facts.P101.manufacturer} CP series)"),
        ("Estimated duration", f"{facts.SOP_TOTAL_MINUTES} minutes"),
        ("Owner", f"{facts.D_KRISHNAN.name}, {facts.D_KRISHNAN.role}"),
        ("Hold points", str(len(hold_points))),
    ))

    s.h1("1.  Purpose")
    s.p(
        "To define the controlled method for replacing the drive-end and non-drive-end rolling "
        f"element bearings on {facts.P101.manufacturer} CP-series horizontal centrifugal pumps, "
        "such that the machine is returned to service within OEM vibration acceptance criteria "
        "and the bearing life stated in the OEM manual is achieved."
    )

    s.h1("2.  Scope and competency")
    s.p(
        f"Applies to {facts.P101.tag}, {facts.P102.tag} and {facts.P105.tag}. Execution requires a "
        "competent person certified on Sulzer CP-series bearing replacement. The certification is "
        "held on the training register; at the date of this revision one technician holds it."
    )

    s.h1("3.  References")
    s.bullets((
        f"{facts.P101.manufacturer} {facts.P101.model} IOM, Section 7.4 — bearing replacement "
        f"criteria ({facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% wear limit).",
        f"{facts.RCA_ID} — root cause analysis of the {facts.INCIDENT_DATE.isoformat()} seizure.",
        "Plant permit-to-work procedure PTW-GEN-002.",
        "Lock-out / tag-out procedure SAF-ELE-006.",
    ))

    s.h1("4.  Personal protective equipment")
    s.bullets((
        "Safety helmet, safety spectacles, steel toe-capped footwear, coverall.",
        "Cut-resistant gloves for bearing handling; heat-resistant gloves for induction fitting.",
        "Hearing protection while the adjacent machine is running.",
    ))

    s.h1("5.  Tools and equipment")
    s.table(
        ["#", "Item"],
        [[str(index + 1), tool] for index, tool in enumerate(facts.SOP_TOOLS)],
        widths_mm=[8, _CONTENT_WIDTH_MM - 8],
    )

    s.h1("6.  Safety notes")
    s.numbered(facts.SOP_SAFETY_NOTES)

    s.page_break()
    s.h1("7.  Procedure")
    s.table(
        ["Step", "Instruction", "Minutes", "Hold point"],
        [
            [str(step.number), step.text, str(step.minutes), "YES" if step.hold_point else ""]
            for step in facts.SOP_STEPS
        ],
        widths_mm=[12, _CONTENT_WIDTH_MM - 46, 16, 18],
        align_right=(2,),
    )
    s.p(
        f"Total estimated duration {facts.SOP_TOTAL_MINUTES} minutes "
        f"({facts.SOP_TOTAL_MINUTES / 60:.0f} hours) excluding permit issue and cool-down. "
        f"{len(hold_points)} hold points require the supervisor's signature before the next step "
        "may start."
    )

    s.h1("8.  Hold points")
    s.table(
        ["Step", "Hold point", "Signatory"],
        [[str(step.number), step.text.split(".")[0] + ".", "Maintenance supervisor"]
         for step in hold_points],
        widths_mm=[12, _CONTENT_WIDTH_MM - 52, 40],
    )

    s.h1("9.  Acceptance criteria")
    s.bullets((
        "Overall vibration below 4.5 mm/s RMS at both bearing housings after 30 minutes running.",
        "Bearing housing temperature stable below 75 degC at rated duty.",
        f"Lube oil header pressure at or above {facts.OEM_LUBE_PRESSURE_MIN_BAR} bar throughout.",
        "Alignment within 0.05 mm rim and 0.03 mm face, recorded on the job card.",
        "No oil leakage from the housings after two hours of operation.",
    ))

    s.h1("10.  Records")
    s.bullets((
        "Completed job card with alignment readings before and after.",
        "Removed bearing wear assessment against the OEM Section 7.4 criteria.",
        "Post-job vibration record with 1x/2x spectrum.",
        "Photographs of the bearing in situ before extraction.",
    ))
    s.small(
        f"Revision history: Rev 1 2018-06-01 initial issue. Rev 2 2020-11-15 induction heating "
        f"mandated. {facts.SOP_REVISION} {facts.SOP_EFFECTIVE_DATE.isoformat()} — hold points "
        f"added and wear assessment against the OEM limit made mandatory following {facts.RCA_ID}."
    )
    return furniture, s


def _regulation(rl: dict[str, Any]) -> tuple[PageFurniture, _Story]:
    """Factory Act 41(b) extract plus the plant compliance note the audit measures against."""
    overdue = facts.days_since(facts.LAST_V201_STATUTORY_INSPECTION) - facts.REGULATION_FREQUENCY_DAYS
    furniture = PageFurniture(
        organisation=facts.PLANT_NAME,
        title=f"{facts.REGULATION_NAME} {facts.REGULATION_CLAUSE} — Extract and Compliance Note",
        document_number="COMP/FA-41B/2023",
        document_date=date(2023, 4, 1),
        classification="STATUTORY COMPLIANCE — CONTROLLED COPY",
    )
    s = _Story(rl)
    s.title(
        f"{facts.REGULATION_NAME} — {facts.REGULATION_CLAUSE}",
        "Statutory extract and plant applicability note  |  Safety, Health and Environment "
        "department",
    )
    s.fields((
        ("Regulation", facts.REGULATION_NAME),
        ("Clause", facts.REGULATION_CLAUSE),
        ("Obligation", facts.REGULATION_OBLIGATION),
        ("Frequency", f"Every {facts.REGULATION_FREQUENCY_DAYS} days (one month)"),
        ("Applies to", ", ".join(facts.REGULATION_APPLIES_TO_TAGS)),
        ("Evidence required", "Signed inspection report by a competent person"),
    ))

    s.h1("1.  Statutory text")
    s.quote(facts.REGULATION_TEXT)

    s.h1("2.  Interpretation for this plant")
    s.p(
        f"The obligation is a {facts.REGULATION_OBLIGATION}: every pressure vessel in the factory "
        f"must be examined by a competent person at intervals not exceeding "
        f"{facts.REGULATION_FREQUENCY_DAYS} days, and the examination must be recorded in the "
        "prescribed register with the competent person's signature and a statement of the vessel's "
        "condition. An examination that is performed but not recorded does not discharge the "
        "obligation. A record without a signature does not discharge it either."
    )

    s.h1("3.  Applicable equipment")
    s.table(
        ["Tag", "Description", "Design pressure", "Statutory class"],
        [
            [facts.V201.tag, facts.V201.name,
             facts.V201.specifications.get("Design pressure", "8.5 barg"),
             f"Pressure vessel — {facts.REGULATION_CLAUSE} applies"],
            [facts.E301.tag, facts.E301.name,
             facts.E301.specifications.get("Shell design pressure", "12 barg"),
             "Heat exchanger — examined under the same register at 12-month intervals"],
        ],
        widths_mm=[20, 56, 28, _CONTENT_WIDTH_MM - 104],
    )

    s.h1("4.  Evidence requirements")
    s.bullets((
        "Inspection report of type 'inspection_report', signed by the competent person.",
        "Date of examination and the condition of the vessel stated explicitly.",
        "Any defect found, and the action taken.",
        "The next examination date, not exceeding one month from the examination.",
        "Entry in Form 8 of the prescribed register, produced on demand to the Inspector.",
    ))

    s.page_break()
    s.h1("5.  Penalty for contravention")
    s.p(facts.REGULATION_PENALTY)

    s.h1("6.  Plant compliance note")
    s.p(
        f"The last recorded statutory examination of {facts.V201.tag} on file is dated "
        f"{facts.LAST_V201_STATUTORY_INSPECTION.isoformat()}, performed by "
        f"{facts.S_RAMASWAMY.name} as the competent person. Against a "
        f"{facts.REGULATION_FREQUENCY_DAYS}-day obligation and a corpus reference date of "
        f"{facts.REFERENCE_DATE.isoformat()}, the next examination fell due on "
        f"{facts.LAST_V201_STATUTORY_INSPECTION.isoformat()} plus one month and no subsequent "
        f"record exists. The examination is therefore {overdue} days overdue."
    )
    s.table(
        ["Examination date", "Competent person", "Finding", "Next due"],
        [
            ["2023-05-12", facts.S_RAMASWAMY.name, "Shell thickness 11.7 mm. No defects.",
             "2023-06-11"],
            ["2023-08-14", facts.S_RAMASWAMY.name, "No defects. PSV-2011 certificate current.",
             "2023-09-13"],
            ["2023-11-20", facts.S_RAMASWAMY.name, "Minor pitting at the manway seat.",
             "2023-12-20"],
            [facts.LAST_V201_STATUTORY_INSPECTION.isoformat(), facts.S_RAMASWAMY.name,
             "Shell thickness 11.6 mm against 12.0 mm nominal.", "2024-03-09"],
            ["Not recorded", "-", "No examination record after "
                                  f"{facts.LAST_V201_STATUTORY_INSPECTION.isoformat()}.",
             "OVERDUE"],
        ],
        widths_mm=[30, 30, 62, _CONTENT_WIDTH_MM - 122],
    )
    s.p(
        f"Note further that {facts.S_RAMASWAMY.name}, the only named competent person for this "
        "vessel on the current register, superannuates on 2024-11-30. Cover for the statutory "
        "examination must be nominated before that date."
    )
    s.small(
        "Controlled copy. Issued by the SHE department. Uncontrolled when printed except for the "
        "copy held in the Unit-2 control room."
    )
    return furniture, s


# ======================================================================================
# Spreadsheet
# ======================================================================================


def _build_maintenance_log() -> bytes:
    """Render the maintenance and condition-monitoring log as a deterministic ``.xlsx``.

    Raises:
        ConfigurationError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - declared in requirements.txt
        raise ConfigurationError(
            "openpyxl is required to generate the maintenance log. Install it with "
            "`pip install openpyxl` (it is already listed in requirements.txt).",
            cause=exc,
        ) from exc

    head_fill = PatternFill("solid", fgColor="26303A")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    body_font = Font(size=10)
    thin = Side(style="thin", color="B0B7BE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    workbook = Workbook()
    workbook.properties.creator = "INDRA demo corpus generator"
    workbook.properties.title = "Maintenance and Condition-Monitoring Log 2022-2024"
    workbook.properties.created = _FIXED_DATETIME
    workbook.properties.modified = _FIXED_DATETIME

    def sheet(name: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              widths: Sequence[int]) -> None:
        ws = workbook.create_sheet(title=name)
        ws.append(list(headers))
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(list(row))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = border
                cell.alignment = wrap
                if isinstance(cell.value, date) and not isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd"
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    readme = workbook.active
    readme.title = "README"
    readme["A1"] = "Maintenance and Condition-Monitoring Log 2022-2024"
    readme["A1"].font = Font(size=14, bold=True)
    notes = [
        ("Plant", facts.PLANT_ADDRESS),
        ("Unit", facts.PLANT_UNIT),
        ("Reference date", facts.REFERENCE_DATE.isoformat()),
        ("Sheets", "Maintenance Records, Condition Readings, Equipment Register, Personnel"),
        ("OEM bearing wear limit", f"{facts.OEM_BEARING_WEAR_LIMIT_PCT:.0f}% "
                                   f"({facts.P101.manufacturer} {facts.P101.model}, Section 7.4)"),
        ("Latest measured wear", f"{facts.MEASURED_BEARING_WEAR_PCT:.0f}% on P-101, "
                                 f"{facts.WORK_ORDER_ID} of {facts.WORK_ORDER_DATE.isoformat()}"),
        ("2022 failure", f"{facts.INCIDENT_ID}, {facts.INCIDENT_FAILURE_MODE}, "
                         f"{facts.INCIDENT_DOWNTIME_HOURS:.0f} h, {facts.INCIDENT_COST_TEXT}"),
        ("Statutory note", f"{facts.REGULATION_CLAUSE} of the {facts.REGULATION_NAME} requires a "
                           f"{facts.REGULATION_OBLIGATION} of {facts.V201.tag}. Last examination "
                           f"on record: {facts.LAST_V201_STATUTORY_INSPECTION.isoformat()}."),
    ]
    for offset, (label, value) in enumerate(notes, start=3):
        readme.cell(row=offset, column=1, value=label).font = Font(bold=True, size=10)
        readme.cell(row=offset, column=2, value=value).font = body_font
        readme.cell(row=offset, column=2).alignment = wrap
    readme.column_dimensions["A"].width = 26
    readme.column_dimensions["B"].width = 96

    sheet(
        "Maintenance Records",
        ("Record ID", "Equipment Tag", "Record Type", "Performed On", "Performed By", "Findings",
         "Action Taken", "Downtime (h)", "Cost (INR)", "Status", "Reference Document"),
        [
            (row.record_id, row.equipment_tag, row.record_type, row.performed_on, row.performed_by,
             row.findings, row.action_taken, row.downtime_hours, row.cost_inr, row.status,
             row.reference_document)
            for row in sorted(facts.MAINTENANCE_ROWS, key=lambda r: (r.performed_on, r.record_id))
        ],
        (14, 14, 13, 13, 16, 62, 52, 12, 13, 10, 30),
    )

    sheet(
        "Condition Readings",
        ("Equipment Tag", "Parameter", "Value", "Unit", "Measured On", "Source Document", "Note"),
        [
            (r.equipment_tag, r.parameter, r.value, r.unit, r.measured_on, r.source_document,
             r.note)
            for r in sorted(facts.READINGS, key=lambda r: (r.measured_on, r.equipment_tag,
                                                          r.parameter))
        ],
        (14, 22, 10, 8, 13, 24, 60),
    )

    sheet(
        "Equipment Register",
        ("Tag", "Name", "Type", "Criticality", "Unit", "Location", "Manufacturer", "Model",
         "Serial Number", "Installed On", "OEM Thresholds", "Notes"),
        [
            (item.tag, item.name, item.equipment_type, item.criticality, item.unit, item.location,
             item.manufacturer or "", item.model or "", item.serial_number or "", item.installed_on,
             "; ".join(f"{key}={value:g}" for key, value in item.oem_thresholds.items()),
             item.notes)
            for item in facts.EQUIPMENT
        ],
        (10, 34, 18, 11, 18, 34, 20, 18, 16, 13, 46, 60),
    )

    sheet(
        "Personnel",
        ("Name", "Employee ID", "Role", "Department", "Years of Service", "Expertise Tags",
         "Superannuation Date", "Documented Contributions", "Contact"),
        [
            (person.name, person.employee_id, person.role, person.department,
             person.years_experience, ", ".join(person.expertise_tags),
             person.retirement_on, person.documented_contributions, person.contact or "")
            for person in facts.PEOPLE
        ],
        (18, 13, 40, 24, 15, 28, 19, 22, 26),
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return _deterministic_zip(buffer.getvalue())


def _deterministic_zip(raw: bytes) -> bytes:
    """Rewrite a ZIP container with fixed entry timestamps and a stable member order.

    ``openpyxl`` stamps wall-clock time into every ZIP entry, so two identical workbooks written a
    second apart differ in bytes and therefore in SHA-256. Content-addressed ingestion (D6) would
    then treat every rehearsal's spreadsheet as a new document.

    Raises:
        BlobStoreError: If the container cannot be rewritten.
    """
    try:
        source = io.BytesIO(raw)
        target = io.BytesIO()
        with zipfile.ZipFile(source) as reader, zipfile.ZipFile(
            target, mode="w", compression=zipfile.ZIP_DEFLATED
        ) as writer:
            for name in sorted(reader.namelist()):
                info = zipfile.ZipInfo(filename=name, date_time=_ZIP_TIMESTAMP)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                writer.writestr(info, reader.read(name))
        return target.getvalue()
    except (OSError, zipfile.BadZipFile, ValueError) as exc:
        raise BlobStoreError(
            "Could not rewrite the generated spreadsheet with deterministic ZIP timestamps. "
            "The workbook itself was built successfully; this is a container-level failure.",
            cause=exc,
        ) from exc


# ======================================================================================
# Generation
# ======================================================================================


def _write_bytes(path: Path, payload: bytes) -> int:
    """Write ``payload`` to ``path``, creating parents. Returns the byte count."""
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(payload)
        return len(payload)
    except OSError as exc:
        raise BlobStoreError(
            f"Could not write {path}. Check that the directory is writable and that the file is "
            "not open in another application.",
            context={"path": str(path)},
            cause=exc,
        ) from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(65536), b""):
                digest.update(block)
    except OSError as exc:
        raise BlobStoreError(
            f"Could not read back {path} to compute its content hash.",
            context={"path": str(path)},
            cause=exc,
        ) from exc
    return digest.hexdigest()


def _record(path: Path, filename: str, *, pages: int | None = None) -> GeneratedFile:
    spec = facts.document_spec(filename)
    return GeneratedFile(
        filename=filename,
        path=path,
        size_bytes=path.stat().st_size,
        sha256=_sha256(path),
        document_type=spec.document_type if spec else "unknown",
        title=spec.title if spec else filename,
        role=spec.role if spec else "corpus",
        pages=pages,
    )


def _png_bytes(image: Any) -> bytes:
    """Encode a PIL image as PNG bytes without a timestamp chunk."""
    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


def generate_corpus(
    *,
    output_dir: Path | None = None,
    seed: int | None = None,
    settings: Settings | None = None,
    skip_existing: bool = False,
) -> CorpusReport:
    """Generate the whole demo corpus and its manifest.

    Args:
        output_dir: Destination. Defaults to ``settings.demo_dir``.
        seed: Base seed for every randomised artefact. Defaults to ``settings.llm_seed``.
        settings: Settings instance. Defaults to :func:`indra.core.config.get_settings`.
        skip_existing: Leave files that are already present untouched. Useful when re-seeding a
            machine where generation has already run; generation is deterministic, so the default
            (regenerate everything) produces identical bytes anyway.

    Returns:
        A :class:`CorpusReport` describing every file written.

    Raises:
        ConfigurationError: If the fact sheet is internally inconsistent, or a required library is
            missing.
        BlobStoreError: If a file cannot be written.
        VisionError: If a drawing cannot be rendered.
    """
    active = settings or get_settings()
    target = Path(output_dir) if output_dir is not None else Path(active.demo_dir)
    base_seed = int(seed if seed is not None else active.llm_seed)

    problems = facts.consistency_report()
    if problems:
        raise ConfigurationError(
            "The demo fact sheet is internally inconsistent, so the generated corpus would make "
            "claims that contradict each other. Fix scripts/demo_facts.py before generating: "
            + " | ".join(problems),
            context={"problem_count": len(problems)},
        )

    target.mkdir(parents=True, exist_ok=True)
    logger.info(
        "generating demo corpus",
        extra={"output_dir": str(target), "seed": base_seed,
               "reference_date": facts.REFERENCE_DATE.isoformat()},
    )

    rl = _reportlab()
    written: list[GeneratedFile] = []
    warnings: list[str] = []

    def exists(filename: str) -> bool:
        candidate = target / filename
        return skip_existing and candidate.exists() and candidate.stat().st_size > 0

    # ---- the handwritten margin annotation, needed before the work order is laid out ----
    annotation = render_handwriting(
        facts.WORK_ORDER_HANDWRITTEN_NOTE, seed=base_seed + _SEED_OFFSET_HANDWRITING
    )
    annotation_png = _png_bytes(annotation)

    pdf_builders: tuple[tuple[str, Any], ...] = (
        (facts.DOC_OEM_MANUAL, lambda: _oem_manual(rl)),
        (facts.DOC_WORK_ORDER, lambda: _work_order(rl, annotation_png=annotation_png)),
        (facts.DOC_INSPECTION, lambda: _inspection(rl)),
        (facts.DOC_SHIFT_LOG, lambda: _shift_log(rl)),
        (facts.DOC_INCIDENT, lambda: _incident(rl)),
        (facts.DOC_RCA, lambda: _rca(rl)),
        (facts.DOC_EMAIL, lambda: _retirement_email(rl)),
        (facts.DOC_SOP, lambda: _sop(rl)),
        (facts.DOC_REGULATION, lambda: _regulation(rl)),
    )

    for filename, builder in pdf_builders:
        path = target / filename
        if exists(filename):
            written.append(_record(path, filename))
            continue
        furniture, story = builder()
        pages = _render_pdf(path, furniture, story, rl)
        written.append(_record(path, filename, pages=pages))
        logger.debug("wrote pdf", extra={"document_filename": filename, "pages": pages})

    # ---- spreadsheet -----------------------------------------------------------------
    log_path = target / facts.DOC_MAINTENANCE_LOG
    if not exists(facts.DOC_MAINTENANCE_LOG):
        _write_bytes(log_path, _build_maintenance_log())
    written.append(_record(log_path, facts.DOC_MAINTENANCE_LOG))

    # ---- drawings --------------------------------------------------------------------
    pid_path = target / facts.DOC_PID
    scan_path = target / facts.DOC_PID_SCANNED
    if not exists(facts.DOC_PID) or not exists(facts.DOC_PID_SCANNED):
        drawing = render_pid()
        save_image(drawing, pid_path)
        save_image(degrade_scan(drawing, seed=base_seed + _SEED_OFFSET_SCAN), scan_path)
    written.append(_record(pid_path, facts.DOC_PID))
    written.append(_record(scan_path, facts.DOC_PID_SCANNED))

    # ---- nameplate photograph (an asset, not part of the ingest corpus) ---------------
    nameplate_path = target / facts.ASSET_NAMEPLATE
    if not exists(facts.ASSET_NAMEPLATE):
        nameplate = render_nameplate(
            (
                ("PUMP TAG", facts.P101.tag),
                ("MODEL", facts.P101.model or ""),
                ("SERIAL No.", facts.P101.serial_number or ""),
                ("RATED FLOW", facts.P101.specifications.get("Rated flow", "")),
                ("RATED HEAD", facts.P101.specifications.get("Rated head", "")),
                ("YEAR", str(facts.P101.installed_on.year)),
            ),
            seed=base_seed + _SEED_OFFSET_NAMEPLATE,
            heading=f"{facts.P101.manufacturer} PUMPS  |  {facts.PLANT_SHORT}",
        )
        save_image(nameplate, nameplate_path, jpeg_quality=84)
    written.append(_record(nameplate_path, facts.ASSET_NAMEPLATE))

    # ---- manifest --------------------------------------------------------------------
    manifest_path = write_manifest(target, tuple(written), seed=base_seed)

    missing = [name for name in facts.CORPUS_FILENAMES
               if not (target / name).exists()]
    if missing:
        warnings.append(f"corpus filenames declared in demo_facts but not generated: {missing}")

    report = CorpusReport(
        output_dir=target,
        files=tuple(written),
        manifest_path=manifest_path,
        seed=base_seed,
        reference_date=facts.REFERENCE_DATE,
        warnings=tuple(warnings),
    )
    logger.info(
        "demo corpus generated",
        extra={"files": len(report.files), "bytes": report.total_bytes,
               "output_dir": str(target)},
    )
    return report


def write_manifest(target: Path, written: Sequence[GeneratedFile], *, seed: int) -> Path:
    """Write ``manifest.json`` next to the corpus and return its path.

    The manifest is the contract between generation and everything downstream: the seeding script
    reads the file list, ``run_demo_check`` reads the cross-document links and the P&ID ground
    truth, and the tests read the fact block instead of hard-coding numbers.
    """
    payload: dict[str, Any] = {
        "schema_version": 1,
        "generator": "scripts.generate_demo_data",
        "seed": seed,
        "reference_date": facts.REFERENCE_DATE.isoformat(),
        "plant": facts.PLANT_ADDRESS,
        "corpus_filenames": list(facts.CORPUS_FILENAMES),
        "files": [item.as_manifest_entry() for item in written],
        "facts": facts.as_manifest_facts(),
        "pid_ground_truth": {
            "image": facts.DOC_PID,
            "scanned_image": facts.DOC_PID_SCANNED,
            "width": PID_WIDTH,
            "height": PID_HEIGHT,
            "symbols": [
                {"tag": symbol.tag, "symbol_class": symbol.symbol_class,
                 "bbox": list(symbol.bbox), "label": symbol.label}
                for symbol in PID_SYMBOLS
            ],
            "connections": [
                {"source": link.source_tag, "target": link.target_tag,
                 "line_type": link.line_type, "pipe_spec": link.pipe_spec}
                for link in PID_CONNECTIONS
            ],
        },
        "summary": list(facts.summary_lines()),
    }
    path = target / MANIFEST_FILENAME
    _write_bytes(path, json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8"))
    return path


def load_manifest(directory: Path | None = None) -> dict[str, Any]:
    """Read a previously written ``manifest.json``.

    Raises:
        ConfigurationError: If the manifest is missing or unreadable — which always means the
            corpus has not been generated yet.
    """
    target = Path(directory) if directory is not None else Path(get_settings().demo_dir)
    path = target / MANIFEST_FILENAME
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ConfigurationError(
            f"No demo manifest at {path}. Generate the corpus first with "
            "`python -m scripts.generate_demo_data`.",
            context={"path": str(path)},
            cause=exc,
        ) from exc
    except (OSError, json.JSONDecodeError) as exc:
        raise ConfigurationError(
            f"The demo manifest at {path} could not be read. Delete it and regenerate the corpus "
            "with `python -m scripts.generate_demo_data`.",
            context={"path": str(path)},
            cause=exc,
        ) from exc


def verify_corpus(directory: Path | None = None) -> tuple[str, ...]:
    """Check a generated corpus on disk against its manifest. Empty tuple means healthy.

    Used by ``scripts/run_demo_check.py`` for the 0:00 upload beat, and by the seeding script
    before it starts an ingest it cannot finish.
    """
    target = Path(directory) if directory is not None else Path(get_settings().demo_dir)
    problems: list[str] = []
    try:
        manifest = load_manifest(target)
    except ConfigurationError as exc:
        return (exc.message,)

    entries = {entry["filename"]: entry for entry in manifest.get("files", [])}
    for filename in facts.CORPUS_FILENAMES:
        if filename not in entries:
            problems.append(f"{filename}: absent from the manifest")
            continue
        path = target / filename
        if not path.exists():
            problems.append(f"{filename}: listed in the manifest but missing from {target}")
            continue
        size = path.stat().st_size
        if size == 0:
            problems.append(f"{filename}: present but empty")
            continue
        expected = int(entries[filename].get("size_bytes", 0))
        if expected and size != expected:
            problems.append(
                f"{filename}: {size} bytes on disk against {expected} in the manifest — the file "
                "has been modified since generation"
            )
        digest = _sha256(path)
        if entries[filename].get("sha256") not in (None, "", digest):
            problems.append(f"{filename}: SHA-256 does not match the manifest")

    if manifest.get("reference_date") != facts.REFERENCE_DATE.isoformat():
        problems.append(
            "manifest reference_date does not match scripts/demo_facts.REFERENCE_DATE — "
            "regenerate the corpus"
        )
    return tuple(problems)


# ======================================================================================
# CLI
# ======================================================================================


def _report(console: Console, report: CorpusReport) -> None:
    console.banner(
        "INDRA — demo corpus generated",
        f"{report.output_dir}   seed={report.seed}   "
        f"reference date={report.reference_date.isoformat()}",
    )
    console.blank()
    console.table(
        ["File", "Type", "Pages", "Bytes", "SHA-256"],
        [
            [item.filename, item.document_type, str(item.pages or "-"),
             f"{item.size_bytes:,}", item.sha256[:12]]
            for item in report.files
        ],
        aligns=["left", "left", "right", "right", "left"],
    )
    console.blank()
    console.kv("Documents for ingestion", str(len(report.corpus_files)))
    console.kv("Total bytes", f"{report.total_bytes:,}")
    console.kv("Manifest", str(report.manifest_path))
    console.blank()
    console.rule("corpus facts")
    for line in facts.summary_lines():
        console.bullet(line)
    console.blank()
    console.rule("cross-document links the demo asserts")
    for link in facts.CROSS_DOCUMENT_LINKS:
        console.bullet(f"{link.link_id}: {link.claim}")
    if report.warnings:
        console.blank()
        for warning in report.warnings:
            console.status("warn", warning)
    console.blank()
    console.status("pass", "corpus ready", "next: python -m scripts.seed_demo_data")


def main(argv: Sequence[str] | None = None) -> int:
    """CLI entry point. Returns a process exit code."""
    parser = argparse.ArgumentParser(
        prog="python -m scripts.generate_demo_data",
        description="Generate INDRA's deterministic synthetic plant corpus.",
    )
    parser.add_argument("--output", type=Path, default=None,
                        help="Destination directory (default: settings.demo_dir).")
    parser.add_argument("--seed", type=int, default=None,
                        help="Base seed for randomised artefacts (default: settings.llm_seed).")
    parser.add_argument("--skip-existing", action="store_true",
                        help="Leave files that already exist untouched.")
    parser.add_argument("--quiet", action="store_true", help="Suppress the report table.")
    args = parser.parse_args(list(argv) if argv is not None else None)

    console = Console()
    try:
        report = generate_corpus(
            output_dir=args.output, seed=args.seed, skip_existing=args.skip_existing
        )
    except IndraError as exc:
        logger.error("corpus generation failed", extra={"error": str(exc)})
        console.blank()
        console.status("fail", "corpus generation failed", exc.message)
        return 1
    if not args.quiet:
        _report(console, report)
    return 0


if __name__ == "__main__":  # pragma: no cover - CLI
    raise SystemExit(main())


__all__ = [
    "MANIFEST_FILENAME",
    "CorpusReport",
    "GeneratedFile",
    "PageFurniture",
    "generate_corpus",
    "load_manifest",
    "main",
    "verify_corpus",
    "write_manifest",
]
